# -*-coding:utf-8-*-
import time
import os
import re
from types import NoneType

import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

path_excel = 'E:\\ExportProject\\TranslateByPython\\translate.xlsx'
from openpyxl.cell.read_only import EmptyCell

# path_excel = 'sample.xlsx'
url_dir_appres = 'E:\\ExportProject\\TranslateByPython'
url_dir_appres_main = 'res'
xml_name = 'arrays_glossary.xml'

ZH = "中文"
ZH_M = "中文繁体"
ZH_HK = "中文繁体_hk"
ZH_TW = "中文繁体_tw"
EN = "英语"
EN_DEFAULT = "英语_default"
KR = "韩语"
DE = "德语"
JP = "日语"
RU = "俄语"
ES = "西班牙语"
IT = "意大利语"
FR = "法语"
VI = "越南语"
PT = "葡萄牙语"


def get_dir_name(language):
    if language == ZH:
        return "values-zh"
    elif language == ZH_HK:
        return "values-zh-rHK"
    elif language == ZH_TW:
        return "values-zh-rTW"
    elif language == EN:
        return "values-en"
    elif language == EN_DEFAULT:
        return "values"
    elif language == KR:
        return "values-ko-rKR"
    elif language == DE:
        return "values-de"
    elif language == JP:
        return "values-ja-rJP"
    elif language == RU:
        return "values-ru"
    elif language == ES:
        return "values-es"
    elif language == IT:
        return "values-it"
    elif language == FR:
        return "values-fr"
    elif language == VI:
        return "values-vi-rVN"
    elif language == PT:
        return "values-pt"
    else:
        return "values"


def save_to_file(str_array, dir_langue, language_title_str):
    # print "保存到文件"
    """
     保存到文件
     :param str_array: string xml的数组\n
     """
    # 如果不存在文件夹
    if not os.path.exists(url_dir_appres):
        os.mkdir(url_dir_appres)

    # 如果不存在文件夹
    dir_main = os.path.join(url_dir_appres, url_dir_appres_main)
    if not os.path.exists(dir_main):
        os.mkdir(dir_main)

    # 如果不存在文件夹
    dir_main_langue = os.path.join(dir_main, dir_langue)
    if not os.path.exists(dir_main_langue):
        os.mkdir(dir_main_langue)

    # 如果存在文件
    path = os.path.join(dir_main_langue, xml_name)
    if os.path.exists(path):
        os.remove(path)
    print  get_dir_name(language_title_str)
    print  "save to file:", path
    fileObject = open(path, 'w')
    for item_str in str_array:
        try:
            fileObject.write(item_str)
            fileObject.write('\n')
        except Exception:
            print  "写入出错=", item_str,
    fileObject.close()


def get_content_format(language_title_str, cell_value_content):
    if (language_title_str == EN or language_title_str == FR or language_title_str == IT) \
            and is_contain_diagonal(cell_value_content):
        cell_value_content = replaceDiagonal(cell_value_content)
    return cell_value_content


def get_arr_str(language_title_str, key, value):
    value = get_content_format(language_title_str, value)
    stringbuffer = "<string-array name= \"" + key;
    stringbuffer = stringbuffer + "\">\n<item>"
    stringbuffer = stringbuffer + value
    stringbuffer = stringbuffer + "</item>\n</string-array>"
    return stringbuffer


def get_item_str(language_title_str, value):
    value = get_content_format(language_title_str, value)
    stringbuffer = "<item>";
    stringbuffer = stringbuffer + value
    stringbuffer = stringbuffer + "</item>"
    return stringbuffer


def getDateTime():
    """
    获得当前时间时间戳
    :return: 时间\n
    """
    now = int(time.time())
    timeStruct = time.localtime(now)
    strTime = time.strftime("%Y_%m_%d_%H_%M_%S", timeStruct)
    return strTime


def replaceDiagonal(str):
    str = str.replace("\'", "\\'")
    str = str.replace("’", "\\'")
    str = str.replace("\\\\", "\\")
    return str


def is_contain_diagonal(str):
    flag = False
    if "\'" in str:
        flag = True
    if "’" in str:
        flag = True
    return flag


def save_to_file_by_language(language_title_str, str_arr):
    if language_title_str == ZH_M:
        save_to_file(str_arr, get_dir_name(ZH_HK), language_title_str)
        save_to_file(str_arr, get_dir_name(ZH_TW), language_title_str)
    elif language_title_str == EN:
        save_to_file(str_arr, get_dir_name(EN), language_title_str)
        save_to_file(str_arr, get_dir_name(EN_DEFAULT), language_title_str)
    else:
        dir_langue = get_dir_name(language_title_str)
        save_to_file(str_arr, dir_langue, language_title_str)


def execl_to_2array(path_excel):
    worksheet = open_execl_file(path_excel)
    if worksheet == None:
        print "worksheet出现错误"
        return None
    ##语言列
    max_column_ = worksheet.max_column
    ## key行
    max_row_ = worksheet.max_row
    # print max_column_, max_row_
    arrListOut = []
    for row in worksheet.rows:
        arrListInner = []
        for cell in row:
            value = cell.value
            if value != None:
                _column = cell.column
                # print "nomal=", cell.value
                arrListInner.append(value.encode("utf-8"))
                if _column == max_column_:
                    arrListOut.append(arrListInner)
            else:
                arrListInner.append("None")
                # print arrListOut
    return arrListOut;


def open_execl_file(path_excel):
    try:
        workbook = load_workbook(path_excel, True)
        worksheet = workbook.worksheets[0]
    except Exception:
        print "文件不存在，无法打开"
        return None
    return worksheet


def is_array(str):
    split_str = str.split("_")
    flag = False
    for item_str in split_str:
        match = is_have_a_number(item_str)
        if match != None:
            flag = True
            break
    return flag


def is_have_a_number(str):
    return re.match(r'[a]\d+', str)


def get_head_array(str):
    split_str = str.split("_")
    arr_length = len(split_str)
    array_head = ""
    for index in range(1, arr_length - 1):
        if (index == 1):
            array_head = split_str[index]
        else:
            array_head = array_head + "_" + split_str[index]
    return array_head


def translate(path_excel):
    array = execl_to_2array(path_excel)
    if array == None:
        print "初使化数组出现错误"
        return
    worksheet = open_execl_file(path_excel)
    ##语言列
    max_column_ = worksheet.max_column
    ## key行
    max_row_ = worksheet.max_row

    array_key_count = len(array)

    array_language = array[0]
    language_count = len(array_language)

    last_head_key_arr = ""
    is_last_head_key_arr = False
    is_add_head_first = False
    print language_count, array_key_count

    for index_language in range(1, language_count):
        print   "---"
        language_title_str = str(array[0][index_language])
        print   "language=", get_dir_name(language_title_str)
        str_arr = []
        is_add_head_first = False;
        for index_key in range(1, array_key_count):
            cell_value_key = array[index_key][0]
            cell_value_content = array[index_key][index_language]
            # print   "key=", cell_value_key, ",value=", cell_value_content
            if not cell_value_key.startswith("ya"):
                if index_key == array_key_count - 1 and len(str_arr) != 0:
                    str_arr.append("</resources>")
                    save_to_file_by_language(language_title_str, str_arr)
                continue
            # print cell_value_key
            if cell_value_content == None:
                print   "cell_value_content None=", cell_value_key,
            if index_key == 1:
                str_arr.append("<?xml version=\"1.0\" encoding=\"uft-8\"?>\n" + "<resources>")
            # 如果是数组存在多个item的话
            flag_is_array = is_array(cell_value_key);
            head_key = get_head_array(cell_value_key)
            if flag_is_array:
                # /*------表示两组多个item的array的头在一起------*/
                if last_head_key_arr == head_key:
                    is_add_head_first = True
                else:
                    is_add_head_first = False
                    if last_head_key_arr != "" and is_last_head_key_arr:
                        str_arr.append("</string-array>")
                if not is_add_head_first:
                    if head_key == ("glossary_osahs_item_3"):
                        str_arr.append("<string-array name= \"" + head_key + "\" formatted=\"false\">");
                    else:
                        str_arr.append("<string-array name= \"" + head_key + "\">");
                    is_add_head_first = True;
                str_arr.append(get_item_str(language_title_str, cell_value_content));
                last_head_key_arr = head_key;
                is_last_head_key_arr = True;
            else:
                # 如果是数组单个item的话
                is_last_head_key_arr = False;
                if is_add_head_first:
                    str_arr.append("</string-array>\n");
                str_arr.append(get_arr_str(language_title_str, head_key, cell_value_content));
                is_add_head_first = False;
            if index_key == array_key_count - 1:
                str_arr.append("</resources>")
                save_to_file_by_language(language_title_str, str_arr)
    print " "
    print language_count - 1, " language", array_key_count - 1, " key"


print "translate start"
translate(path_excel)
print " "
input("translate end")
