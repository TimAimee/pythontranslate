# -*-coding:utf-8-*-
import time
import os
from types import NoneType

import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

path_excel = 'E:\\ExportProject\\TranslateByPython\\translate.xlsx'
from openpyxl.cell.read_only import EmptyCell

# path_excel = 'sample.xlsx'
url_dir_appres = 'E:\\ExportProject\\TranslateByPython'
url_dir_appres_main = 'res'
xml_name = 'string.xml'

ZH = "中文"
ZH_M = "中文繁体"
ZH_HK = "中文繁体_hk"
ZH_TW = "中文繁体_tw"
EN = "英语"
EN_DEFAULT = "英语_default"
KR = "韩语"
DE = "德语"
JP = "日语"
RU = "俄语"
ES = "西班牙语"
IT = "意大利语"
FR = "法语"
VI = "越南语"
PT = "葡萄牙语"


def get_dir_name(language):
    if language == ZH:
        return "values-zh"
    elif language == ZH_HK:
        return "values-zh-rHK"
    elif language == ZH_TW:
        return "values-zh-rTW"
    elif language == EN:
        return "values-en"
    elif language == EN_DEFAULT:
        return "values"
    elif language == KR:
        return "values-ko-rKR"
    elif language == DE:
        return "values-de"
    elif language == JP:
        return "values-ja-rJP"
    elif language == RU:
        return "values-ru"
    elif language == ES:
        return "values-es"
    elif language == IT:
        return "values-it"
    elif language == FR:
        return "values-fr"
    elif language == VI:
        return "values-vi-rVN"
    elif language == PT:
        return "values-pt"
    else:
        return "values"


def save_to_file(str_array, dir_langue, language_title_str):
    # print "保存到文件"
    """
     保存到文件
     :param str_array: string xml的数组\n
     """
    # 如果不存在文件夹
    if not os.path.exists(url_dir_appres):
        os.mkdir(url_dir_appres)

    # 如果不存在文件夹
    dir_main = os.path.join(url_dir_appres, url_dir_appres_main)
    if not os.path.exists(dir_main):
        os.mkdir(dir_main)

    # 如果不存在文件夹
    dir_main_langue = os.path.join(dir_main, dir_langue)
    if not os.path.exists(dir_main_langue):
        os.mkdir(dir_main_langue)

    # 如果存在文件
    path = os.path.join(dir_main_langue, xml_name)
    if os.path.exists(path):
        os.remove(path)
    print  get_dir_name(language_title_str)
    print  "save to file:", path
    fileObject = open(path, 'w')
    for item_str in str_array:
        try:
            fileObject.write(item_str)
            fileObject.write('\n')
        except Exception:
            print  "写入出错=", item_str,
    fileObject.close()


def get_str_by_key_value(key, value):
    stringbuffer = "<string name=\"" + key;
    if key == "tip":
        stringbuffer = stringbuffer + "\" formatted=\"false\">"
    else:
        stringbuffer = stringbuffer + "\">"
    stringbuffer = stringbuffer + value
    stringbuffer = stringbuffer + "</string>"
    return stringbuffer


def getDateTime():
    """
    获得当前时间时间戳
    :return: 时间\n
    """
    now = int(time.time())
    timeStruct = time.localtime(now)
    strTime = time.strftime("%Y_%m_%d_%H_%M_%S", timeStruct)
    return strTime


def replaceDiagonal(str):
    str = str.replace("\'", "\\'")
    str = str.replace("’", "\\'")
    str = str.replace("\\\\", "\\")
    return str


def is_contain_diagonal(str):
    flag = False
    if "\'" in str:
        flag = True
    if "’" in str:
        flag = True
    return flag


def save_to_file_by_language(language_title_str, str_arr):
    if language_title_str == ZH_M:
        save_to_file(str_arr, get_dir_name(ZH_HK), language_title_str)
        save_to_file(str_arr, get_dir_name(ZH_TW), language_title_str)
    elif language_title_str == EN:
        save_to_file(str_arr, get_dir_name(EN), language_title_str)
        save_to_file(str_arr, get_dir_name(EN_DEFAULT), language_title_str)
    else:
        dir_langue = get_dir_name(language_title_str)
        save_to_file(str_arr, dir_langue, language_title_str)


def execl_to_2array(path_excel):
    worksheet = open_execl_file(path_excel)
    if worksheet == None:
        print "worksheet出现错误"
        return None
    ##语言列
    max_column_ = worksheet.max_column
    ## key行
    max_row_ = worksheet.max_row
    # print max_column_, max_row_
    arrListOut = []
    for row in worksheet.rows:
        arrListInner = []
        for cell in row:
            value = cell.value
            if value != None:
                _column = cell.column
                # print "nomal=", cell.value
                arrListInner.append(value.encode("utf-8"))
                if _column == max_column_:
                    arrListOut.append(arrListInner)
            else:
                arrListInner.append("None")
                # print arrListOut
    return arrListOut;


def open_execl_file(path_excel):
    try:
        workbook = load_workbook(path_excel, True)
        worksheet = workbook.worksheets[0]
    except Exception:
        print "文件不存在，无法打开"
        return None
    return worksheet


def translate(path_excel):
    array = execl_to_2array(path_excel)
    if array == None:
        print "初使化数组出现错误"
        return
    worksheet = open_execl_file(path_excel)
    ##语言列
    max_column_ = worksheet.max_column
    ## key行
    max_row_ = worksheet.max_row

    array_key_count = len(array)

    array_language = array[0]
    language_count = len(array_language)

    print language_count, array_key_count

    for index_language in range(1, language_count):
        print   "---"
        language_title_str = str(array[0][index_language])
        print   "language=", get_dir_name(language_title_str)
        str_arr = []
        for index_key in range(1, array_key_count):
            cell_value_key = array[index_key][0]
            cell_value_content = array[index_key][index_language]
            # print   "key=", cell_value_key, ",value=", cell_value_content
            if cell_value_key == None or "None" in cell_value_key or cell_value_key.startswith("ya"):
                if index_key == array_key_count - 1:
                    str_arr.append("</resources>")
                    save_to_file_by_language(language_title_str, str_arr)
                continue
            if cell_value_content == None:
                print   "cell_value_content None=", cell_value_key,
            if index_key == 1:
                str_arr.append("<?xml version=\"1.0\" encoding=\"uft-8\"?>\n" + "<resources>")
            if (language_title_str == EN or language_title_str == FR or language_title_str == IT) \
                    and is_contain_diagonal(cell_value_content):
                cell_value_content = replaceDiagonal(cell_value_content)
            try:
                value = get_str_by_key_value(cell_value_key, cell_value_content)
            except Exception:
                print "----->Error,key=", cell_value_key, ",value=", cell_value_content
            str_arr.append(value)
            if index_key == array_key_count - 1:
                str_arr.append("</resources>")
                save_to_file_by_language(language_title_str, str_arr)
    print " "
    print language_count - 1, " language", array_key_count - 1, " key"


print "translate start"
translate(path_excel)
print " "
input("translate end")
