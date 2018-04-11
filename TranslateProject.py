# -*-coding:utf-8-*-
import time
import os
import re
import sys
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook


class Translate(object):
    path_excel = 'E:\\Public_resorces\\HBand_translations\\trunk\\HBandTranslate.xlsx'
    url_dir_appres = 'E:\\ExportProject\\TranslateByPython'
    url_dir_appres_main = 'res'
    ZH = '中文'
    ZH_M = "中文繁体"
    ZH_HK = "中文繁体_hk"
    ZH_TW = "中文繁体_tw"
    EN = "英语"
    EN_DEFAULT = "英语_default"
    KR = "韩语"
    DE = "德语"
    JP = "日语"
    RU = "俄语"
    ES = "西班牙语"
    IT = "意大利语"
    FR = "法语"
    VI = "越南语"
    PT = "葡萄牙语"

    def get_datetime(self):
        """获得当前格式化的时间戳

        :return: 当前格式化的时间戳
        """
        now = int(time.time())
        timeStruct = time.localtime(now)
        strTime = time.strftime("%Y_%m_%d_%H_%M_%S", timeStruct)
        return strTime

    def _set_filexmlname(self):
        """设置xml保存时的文件名

        :return: xml文件名
        """
        return None

    def _get_translate_str(self, array):
        """根据二维数组得到翻译的结果"""
        return None

    def _change_content(self, cell_value_content, language_title_str):
        """针对国家语言来替换存在引号的翻译内容

        :param cell_value_content: 翻译内容
        :param language_title_str:  国家语言
        :return:
        """
        if (language_title_str == self.EN or language_title_str == self.FR or language_title_str == self.IT) \
                and self._is_contain_diagonal(cell_value_content):
            cell_value_content = self._replace_diagonal(cell_value_content)
        return cell_value_content

    def _replace_diagonal(self, str):
        """替换存在引号的翻译内容 ,将翻译内容中存在 \'、'、’的字符统一转换成 \'

        :param str: 目标字符串
        :return:  替换后的字符串
        """
        str = str.replace("\'", "\\'")
        str = str.replace("’", "\\'")
        str = str.replace("\\\\", "\\")
        return str

    def _is_contain_diagonal(self, str):
        """判断目标字符串是否包含引号

        :param str: 目标字符串
        :return: 'True'表示存在，'False'表示不存在
        """
        flag = False
        if "\'" in str:
            flag = True
        if "’" in str:
            flag = True
        return flag

    def do_translate(self, array2):
        """提供一个二维数组进行翻译

        :param array2: 二维数组
        """
        list_transtlate_dic = self._get_translate_str(array2);
        filename = self._set_filexmlname();
        self.__save_file(filename, list_transtlate_dic)

    def __save_file(self, filename, list_transtlate_dic):
        """保存翻译的内容到相应的文件上

        :param filename: 文件名
        :param list_transtlate_dic:字典列表的翻译结果
        :return:
        """
        if list_transtlate_dic == None or len(list_transtlate_dic) == 0: return
        for transtlate_dic in list_transtlate_dic:
            print "***"
            self.__save_xml_by_language(filename, transtlate_dic["result"], transtlate_dic["language"])

    def execl_to_2array(self):
        """ 打开指定文件路径的excel,并读取excel的行列数据，生成一个二维数组

        :return: 二维数组
        """
        worksheet = self.__open_execl_file(self.path_excel)
        if worksheet == None:
            print "worksheet出现错误"
            return None
        ##语言列
        max_column_ = worksheet.max_column
        ## key行
        max_row_ = worksheet.max_row
        # print max_column_, max_row_
        arrListOut = []
        for row in worksheet.rows:
            arrListInner = []
            for cell in row:
                value = cell.value
                if value != None:
                    _column = cell.column
                    # print "nomal=", cell.value
                    arrListInner.append(value.encode("utf-8"))
                    if _column == max_column_:
                        arrListOut.append(arrListInner)
                else:
                    arrListInner.append("None")
                    # print arrListOut
        return arrListOut;

    def __open_execl_file(self, path_excel):
        """打开相应文件路径的excel

        :param path_excel: 文件路径
        :return: workbook对象
        """
        try:
            workbook = load_workbook(path_excel, True)
            worksheet = workbook.worksheets[0]
        except Exception:
            print "文件不存在，无法打开"
            return None
        return worksheet

    def __get_dir_name(self, language):
        """根据国家语言返回对应的文件夹名字，如果"意大利语"->"values-it"

        :param language: 国家语言
        :return: 文件夹名字
        """
        if language == self.ZH:
            return "values-zh"
        elif language == self.ZH_HK:
            return "values-zh-rHK"
        elif language == self.ZH_TW:
            return "values-zh-rTW"
        elif language == self.EN:
            return "values-en"
        elif language == self.EN_DEFAULT:
            return "values"
        elif language == self.KR:
            return "values-ko-rKR"
        elif language == self.DE:
            return "values-de"
        elif language == self.JP:
            return "values-ja-rJP"
        elif language == self.RU:
            return "values-ru"
        elif language == self.ES:
            return "values-es"
        elif language == self.IT:
            return "values-it"
        elif language == self.FR:
            return "values-fr"
        elif language == self.VI:
            return "values-vi-rVN"
        elif language == self.PT:
            return "values-pt"
        else:
            return "values"

    def __save_xml_by_language(self, filename, result, language_title_str):
        """根据国家语言来选择保存的不同情况，其中英语以及繁体要保存两次

        :param filename: 文件名,如"string.xml"
        :param result: 列表类型，翻译的结果
        :param language_title_str: 国家语言
        """
        if language_title_str == self.ZH_M:
            self.__save_to_file(filename, result, self.ZH_HK)
            self.__save_to_file(filename, result, self.ZH_TW)
        elif language_title_str == self.EN:
            self.__save_to_file(filename, result, self.EN)
            self.__save_to_file(filename, result, self.EN_DEFAULT)
        else:
            self.__save_to_file(filename, result, language_title_str)

    def __save_to_file(self, filename, result, language_title_str):
        """保存内容到文件

        :param filename: 文件名,如"string.xml"
        :param result: 列表类型，翻译的结果
        :param language_title_str: 国家语言
        """
        dir_language = self.__get_dir_name(language_title_str)
        # print "保存到文件"
        """
         保存到文件
         :param result: string xml的数组\n
         """
        # 如果不存在文件夹
        if not os.path.exists(self.url_dir_appres):
            os.mkdir(self.url_dir_appres)

        # 如果不存在文件夹
        dir_main = os.path.join(self.url_dir_appres, self.url_dir_appres_main)
        if not os.path.exists(dir_main):
            os.mkdir(dir_main)

        # 如果不存在文件夹
        dir_main_langue = os.path.join(dir_main, dir_language)
        if not os.path.exists(dir_main_langue):
            os.mkdir(dir_main_langue)

        # 如果存在文件
        path = os.path.join(dir_main_langue, filename)
        if os.path.exists(path):
            os.remove(path)
        print  dir_language
        print  "save to file:", path
        fileObject = open(path, 'w')
        for item_str in result:
            try:
                fileObject.write(item_str)
                fileObject.write('\n')
            except Exception:
                print  "写入出错=", item_str,
        fileObject.close()

"""
用于翻译普通string的类
"""
class TranslateStr(Translate):
    __xml_name_str = 'string.xml'

    def _get_translate_str(self, array):
        """根据二维数组得到翻译的结果

            :param array: 二维数组
            :return : 字典列表,[ {"language": language_title_str, "result": result_arr}, {"language": language_title_str, "result": result_arr}]
            ，其中result_arr就是由对应的翻译结果的得到的xml数组，后期保存的xml数据使用的就是它
        """
        array_key_count = len(array)
        array_language = array[0]
        language_count = len(array_language)
        list_transtlate_dic = []
        # print language_count, array_key_count
        for index_language in range(2, language_count):
            language_title_str = str(array[0][index_language])
            # print  "language=", language_title_str, self._get_dir_name(language_title_str),
            result_arr = []
            for index_key in range(1, array_key_count):
                cell_value_key = array[index_key][1]
                cell_value_content = array[index_key][index_language]
                if index_key == 1:
                    result_arr.append("<?xml version=\"1.0\" encoding=\"uft-8\"?>\n" + "<resources>")
                # print   "key=", cell_value_key, ",value=", cell_value_content
                result = self.__get_result(cell_value_content, cell_value_key, language_title_str)
                if result != None:
                    result_arr.append(result);
                if index_key == array_key_count - 1:
                    result_arr.append("</resources>")
                    transtlate_dic = {"language": language_title_str, "result": result_arr}
                    list_transtlate_dic.append(transtlate_dic)
        print "--------------------"
        print language_count - 1, " language", array_key_count - 1, " key"
        print "--------------------"
        return list_transtlate_dic

    def __get_result(self, cell_value_content, cell_value_key, language_title_str):
        """根据翻译key内容，翻译content内容，国家语言，返回相应Android格式的字符串结果

        :param cell_value_content: 翻译content内容
        :param cell_value_key: 翻译key内容
        :param language_title_str: 国家语言
        :return: Android格式的字符串结果
        """
        if cell_value_key == None or "None" in cell_value_key or cell_value_key.startswith("ya"):
            return None
        else:
            if cell_value_content == None:
                print   "cell_value_content None=", cell_value_key,
            cell_value_content = self._change_content(cell_value_content, language_title_str)
            try:
                value = self.__get_str_by_key_value(cell_value_key, cell_value_content)
            except Exception:
                print "----->Error,key=", cell_value_key, ",value=", cell_value_content
            return value

    def _set_filexmlname(self):
        """设置xml保存时的文件名

        :return: xml文件名
        """
        return self.__xml_name_str

    def __get_str_by_key_value(self, cell_value_key, cell_value_content):
        """返回根据content，返回Android格式的内容,
             如'key，content'-->'<string name="key">content</string>'

             :param key: 翻译key内容
             :param value: 翻译value内容
             :return: Android格式的内容
             """
        value = "\t<string name=\"" + cell_value_key;
        if cell_value_key == "tip":
            value = value + "\" formatted=\"false\">"
        else:
            value = value + "\">"
        value = value + cell_value_content
        value = value + "</string>"
        return value


"""
用于翻译ios的类
"""

class TranslateIOS(Translate):
    __xml_name_str = 'string_ios.xml'

    def _get_translate_str(self, array):
        """根据二维数组得到翻译的结果

        :param array: 二维数组
        :return : 字典列表,[ {"language": language_title_str, "result": result_arr}, {"language": language_title_str, "result": result_arr}]
        ，其中result_arr就是由对应的翻译结果的得到的xml数组，后期保存的xml数据使用的就是它
        """
        array_key_count = len(array)
        array_language = array[0]
        language_count = len(array_language)
        list_transtlate_dic = []
        # print language_count, array_key_count
        for index_language in range(2, language_count):
            language_title_str = str(array[0][index_language])
            # print  "language=", language_title_str, self._get_dir_name(language_title_str),
            result_arr = []
            for index_key in range(1, array_key_count):
                cell_value_key = array[index_key][0]
                cell_value_content = array[index_key][index_language]
                # print   "key=", cell_value_key, ",value=", cell_value_content
                result = self.__get_result(cell_value_content, cell_value_key, language_title_str)
                if result != None:
                    result_arr.append(result);
                if index_key == array_key_count - 1:
                    transtlate_dic = {"language": language_title_str, "result": result_arr}
                    list_transtlate_dic.append(transtlate_dic)
        print "--------------------"
        print language_count - 1, " language", array_key_count - 1, " key"
        print "--------------------"
        return list_transtlate_dic

    def __get_result(self, cell_value_content, cell_value_key, language_title_str):
        """根据翻译key内容，翻译content内容，国家语言，返回相应Android格式的字符串结果

        :param cell_value_content: 翻译content内容
        :param cell_value_key: 翻译key内容
        :param language_title_str: 国家语言
        :return: Android格式的字符串结果
        """
        if cell_value_key == None or "None" in cell_value_key or cell_value_key.startswith("ya"):
            return None
        else:
            if cell_value_content == None:
                print   "cell_value_content None=", cell_value_key,
            cell_value_content = self._change_content(cell_value_content, language_title_str)
            try:
                value = self.__get_str_by_key_value(cell_value_key, cell_value_content)
            except Exception:
                print "----->Error,key=", cell_value_key, ",value=", cell_value_content
            return value

    def _set_filexmlname(self):
        """设置xml保存时的文件名

        :return: xml文件名
        """
        return self.__xml_name_str

    def __get_str_by_key_value(self, cell_value_key, cell_value_content):
        """返回根据content，返回Android格式的内容,
        如'key，content'-->'"key" = "content"'

        :param key: 翻译key内容
        :param content: 翻译content内容
        :return: Android格式的内容
        """
        value = "\"" + cell_value_key + "\" = " + "\"" + cell_value_content + "\""
        return value


"""
用于翻译Android中string-array的类
"""


class TranslateArr(Translate):
    xml_name_array = 'arrays_glossary.xml'

    def _get_translate_str(self, array):
        """根据二维数组得到翻译的结果

        :param array: 二维数组
        :return : 字典列表,[ {"language": language_title_str, "result": result_arr}, {"language": language_title_str, "result": result_arr}]
        ，其中result_arr就是由对应的翻译结果的得到的xml数组，后期保存的xml数据使用的就是它
        """
        array_key_count = len(array)
        array_language = array[0]
        language_count = len(array_language)
        list_transtlate_dic = []
        # print language_count, array_key_count
        for index_language in range(2, language_count):
            language_title_str = str(array[0][index_language])
            # print  "language=", language_title_str, self._get_dir_name(language_title_str),
            result_arr = []
            for index_key in range(1, array_key_count):
                cell_value_key = array[index_key][1]
                cell_value_content = array[index_key][index_language]
                if index_key == 1:
                    result_arr.append("<?xml version=\"1.0\" encoding=\"uft-8\"?>\n" + "<resources>")
                # print   "key=", cell_value_key, ",value=", cell_value_content
                result = self.__get_result(cell_value_content, cell_value_key, language_title_str)
                if result != None:
                    result_arr.append(result)
                if index_key == array_key_count - 1:
                    result_arr.append("</resources>")
                    transtlate_dic = {"language": language_title_str, "result": result_arr}
                    list_transtlate_dic.append(transtlate_dic)
        print "--------------------"
        print language_count - 1, " language", array_key_count - 1, " key"
        print "--------------------"
        return list_transtlate_dic

    def __get_result(self, cell_value_content, cell_value_key, language_title_str):
        """根据翻译key内容，翻译content内容，国家语言，返回相应Android格式的字符串结果

        :param cell_value_key: 翻译key内容
        :param cell_value_content: 翻译content内容
        :param language_title_str: 国家语言
        :return: Android格式的字符串结果
        """
        if cell_value_key == None or "None" in cell_value_key or not cell_value_key.startswith("ya"):
            return None
        if cell_value_content == None:
            print   "cell_value_content None=", cell_value_key,
        cell_value_content = self._change_content(cell_value_content, language_title_str)
        if self.__is_array(cell_value_key):
            result = self.__get_result_array(cell_value_key, cell_value_content)
        else:
            result = self.__get_result_single(cell_value_key, cell_value_content)
        return result

    def __get_result_array(self, cell_value_key, cell_value_content):
        """根据翻译key内容，翻译content内容，国家语言，返回相应Android格式的内容

        :param cell_value_key: 翻译key内容
        :param cell_value_content: 翻译content内容
        :return: Android格式的内容
        """
        cell_value_key_really = self.__get_head_key_array(cell_value_key)
        item_str = self.__get_item_str(cell_value_content)
        if self.__is_exits_start(cell_value_key):
            value = "\t<string-array name= \"" + cell_value_key_really + "\">\n" + item_str
        elif self.__is_exits_middle(cell_value_key):
            value = item_str
        elif self.__is_exits_stop(cell_value_key):
            value = item_str + "\n\t</string-array>"
        return value

    def __get_result_single(self, cell_value_key, cell_value_content):
        """根据翻译key内容，翻译content内容，返回相应Android格式的内容

        :param cell_value_key: 翻译key内容
        :param cell_value_content: 翻译content内容
        :return: Android格式的内容
        """
        cell_value_key_really = self.__get_head_key_item(cell_value_key)
        try:
            value = self.__get_str_by_key_value(cell_value_key_really, cell_value_content)
        except Exception:
            print "----->Error,key=", cell_value_key, ",value=", cell_value_content
        return value

    def __get_str_by_key_value(self, cell_value_key, cell_value_content):
        """返回根据key以及content，返回Android格式的内容,
        如\n'content'-->'<string-array name="key"><item>content</item></string-array>'

        :param cell_value_key: 翻译key内容
        :param cell_value_content: 翻译content内容
        :return: Android格式的内容
        """
        value = "\t<string-array name= \"" + cell_value_key;
        value = value + "\">\n\t\t<item>"
        value = value + cell_value_content
        value = value + "</item>\n\t</string-array>"
        return value

    def __get_item_str(self, cell_value_content):
        """返回根据content，返回Android格式的内容,
        如'content'-->'<item>content</item>'

        :param cell_value_content: 翻译content内容
        :return: Android格式的内容
        """
        value = "\t\t<item>";
        value = value + cell_value_content
        value = value + "</item>"
        return value

    def __get_head_key_array(self, str):
        """返回array_array的真实key,如'ya_key_a1_start'-->'key'

        :param str: excel中的key
        :return: 真实的key
        """
        split_str = str.split("_")
        arr_length = len(split_str)
        array_head = ""
        for index in range(1, arr_length - 2):
            if (index == 1):
                array_head = split_str[index]
            else:
                array_head = array_head + "_" + split_str[index]
        return array_head

    def __get_head_key_item(self, str):
        """返回array_item的真实key,如'ya_key'-->'key'

        :param str: excel中的key
        :return: 真实的key
        """
        split_str = str.split("_")
        arr_length = len(split_str)
        array_head = ""
        for index in range(1, arr_length):
            if (index == 1):
                array_head = split_str[index]
            else:
                array_head = array_head + "_" + split_str[index]
        return array_head

    def __is_array(self, str):
        """字符串是否含有'a'+数字的组合字样,如"a3"返回True

            :param str: 目标字符串
            :return: 'True'表示存在，'False'表示不存在
            """
        return self.__is_array_match(r'[a]\d+', str)

    def __is_exits_start(self, str):
        """字符串是否含有'start'字样，如"a_start"返回True

            :param str: 目标字符串
            :return: 'True'表示存在，'False'表示不存在
            """
        return self.__is_array_match(r'start', str)

    def __is_exits_middle(self, str):
        """字符串是否含有'middle'字样,如"a_middle"返回True

            :param str: 目标字符串
            :return: 'True'表示存在，'False'表示不存在
            """
        return self.__is_array_match(r'middle', str)

    def __is_exits_stop(self, str):
        """字符串是否含有'stop'字样,如"a_stop"返回True

        :param str: 目标字符串
        :return: 'True'表示存在，'False'表示不存在
        """
        return self.__is_array_match(r'stop', str)

    def __is_array_match(self, format_str, str):
        """规则字符串的匹配

        :param format_str: 匹配标准
        :param str: 目标字符串
        :return: 'True'表示匹配成功，'False'表示匹配失败
        """
        split_str = str.split("_")
        flag = False
        for item_str in split_str:
            match = re.match(format_str, item_str)
            if match != None:
                flag = True
                break
        return flag

    def _set_filexmlname(self):
        """设置xml保存时的文件名

        :return: xml文件名
        """
        return self.xml_name_array


def __select_translate(selet, array2):
    """根据对应的选择，执行相应的翻译步骤

     :param selet: 用户的选择.
     :param array2: 读取excel后转换的二维数组
     :return string: 选择的描述.
    """
    select_str = ""
    if selet == '1' or selet == '3':
        select_str = "translate  string"
        translate_str_obj = TranslateStr()
        translate_str_obj.do_translate(array2)
    if selet == '2' or selet == '3':
        select_str = "translate  array"
        translate_arr_obj = TranslateArr()
        translate_arr_obj.do_translate(array2)
    if selet == '3':
        select_str = "translate string && array"
    if selet == '4':
        select_str = "translate  ios"
        translate_ios_obj = TranslateIOS()
        translate_ios_obj.do_translate(array2)
    return select_str


if '__main__' == __name__:
    """将一定规则的翻译好的execl，转换成Android/IOS需要的翻译文件"""

    selet = '1'
    selet = sys.argv[1]
    translateObj = Translate()

    start_time = translateObj.get_datetime()

    array2 = translateObj.execl_to_2array()
    select_str = __select_translate(selet, array2)

    end_time = translateObj.get_datetime()
    print ""
    print select_str
    print "translate  start time:", start_time
    print "translate  ender time:", end_time
    print ""
    print "----------TimAimee----------"
    print ""
    print ""
    print "-remember happy everyday!!!-"
    print ""
    print ""
    print "----------------------------"
    raw_input("")
